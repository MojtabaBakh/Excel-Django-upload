
from django.shortcuts import render,HttpResponse,redirect
import openpyxl
from . import models
from django.contrib.auth.decorators import login_required
from django.contrib import admin
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test


def home(request):
    return render(request, 'myapp/home.html')
    


@login_required(login_url = "/accounts/login")
def showdata(request):
    username = None
    if request.user.is_authenticated and  int(request.user.get_username()) :
        username = request.user.get_username()
        print(username)
        thisdata = models.Peoplesdata.objects.all()
        print(thisdata)
        senddata=models.Peoplesdata 
        for anydata in thisdata :
            if anydata.kodemeli == int(username) :

                senddata=anydata
                senddata.ozviyat=200000*senddata.tedadhesab
                senddata.moavagemahiyane=senddata.pardakhti-(senddata.mablaghghest+senddata.ozviyat)
                senddata.mandesincemahegabl=senddata.mandeghabli-(senddata.pardakhti-senddata.ozviyat)

        print(senddata.kodemeli)
        args = {'senddata':senddata}
        return render(request, 'myapp/showdata.html', args)
        print( senddata.kokodemeli)
        print( senddata.groupnumber)
    else :
        return redirect ('myapp:list')


  

@user_passes_test(lambda u: u.is_superuser)
def savedata(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        # print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        # print(active_sheet)

        # reading a cell
        # print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        models.Peoplesdata.objects.all().delete()
        for row in worksheet.iter_rows():
            
            
            row_data = list()
            my_data = models.Peoplesdata()
            my_data.radif=row[0].value
            my_data.groupnumber=row[1].value
            my_data.fullname=row[2].value
            my_data.tedadhesab=row[3].value
            my_data.mandeghabli=row[4].value
            my_data.mablaghghest=row[5].value
            my_data.pardakhti=row[6].value
            my_data.Tozihat =row[7].value
            my_data.kodemeli=row[8].value
            my_data.save()
            # for cell in row:

            #     row_data.append(str(cell.value))
            #     # print(cell.value)
            # excel_data.append(row_data)

        return render( request, 'myapp/index.html', {"excel_data":excel_data})








